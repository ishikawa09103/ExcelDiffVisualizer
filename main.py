import streamlit as st
import pandas as pd
from st_aggrid import AgGrid, GridOptionsBuilder
import comparison
import utils
import styles
from utils import get_excel_cell_reference, get_excel_range_reference
from openpyxl import load_workbook
import tempfile
import os

st.set_page_config(
    page_title="Excel Comparison Tool",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Apply custom CSS
styles.apply_custom_css()

def main():
    try:
        st.title("Excel ファイル比較ツール")
        
        # Initialize state management
        if 'upload_error' not in st.session_state:
            st.session_state.upload_error = None
        if 'comparison_error' not in st.session_state:
            st.session_state.comparison_error = None
        
        # File upload section with error handling
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("ファイル 1")
            try:
                file1 = st.file_uploader(
                    "1つ目のExcelファイルをアップロード",
                    type=['xlsx', 'xls'],
                    key="file1_uploader",
                    on_change=lambda: setattr(st.session_state, 'upload_error', None)
                )
            except Exception as e:
                st.error(f"ファイル1のアップロード中にエラーが発生しました: {str(e)}")
                file1 = None
        
        with col2:
            st.subheader("ファイル 2")
            try:
                file2 = st.file_uploader(
                    "2つ目のExcelファイルをアップロード",
                    type=['xlsx', 'xls'],
                    key="file2_uploader",
                    on_change=lambda: setattr(st.session_state, 'upload_error', None)
                )
            except Exception as e:
                st.error(f"ファイル2のアップロード中にエラーが発生しました: {str(e)}")
                file2 = None

        # Display any previous upload errors
        if st.session_state.upload_error:
            st.error(st.session_state.upload_error)
            st.session_state.upload_error = None

        if file1 and file2:
            try:
                # Save uploaded files to temporary locations
                try:
                    temp_dir = tempfile.mkdtemp()
                    file1_path = os.path.join(temp_dir, "file1.xlsx")
                    file2_path = os.path.join(temp_dir, "file2.xlsx")
                    
                    with open(file1_path, 'wb') as f:
                        f.write(file1.getvalue())
                    with open(file2_path, 'wb') as f:
                        f.write(file2.getvalue())
                    
                    # Get workbook information using openpyxl
                    try:
                        wb1 = load_workbook(file1_path)
                        wb2 = load_workbook(file2_path)
                        
                        # シート名の取得
                        sheets1 = wb1.sheetnames
                        sheets2 = wb2.sheetnames
                        
                        # クリーンアップ
                        wb1.close()
                        wb2.close()
                    except Exception as e:
                        st.error(f"シート情報の取得中にエラーが発生しました: {str(e)}")
                        return
                    
                    if not sheets1 or not sheets2:
                        st.error("有効なシートが見つかりません。")
                        return

                    # Debug output for sheet information
                    st.write(f"ファイル1のシート数: {len(sheets1)}")
                    st.write(f"ファイル1のシート名: {', '.join(sheets1)}")
                    st.write(f"ファイル2のシート数: {len(sheets2)}")
                    st.write(f"ファイル2のシート名: {', '.join(sheets2)}")
                    
                    # シート選択の前にシートの追加/削除を確認
                    added_sheets = set(sheets2) - set(sheets1)
                    deleted_sheets = set(sheets1) - set(sheets2)

                    if added_sheets:
                        st.info(f"追加されたシート: {', '.join(added_sheets)}")
                    if deleted_sheets:
                        st.info(f"削除されたシート: {', '.join(deleted_sheets)}")

                    # シート選択のロジックを変更（シート数チェックを削除）
                    st.subheader("シートの選択")
                    common_sheets = list(set(sheets1) & set(sheets2))
                    if common_sheets:
                        st.write("共通するシートの比較:")
                        selected_sheets1 = st.multiselect("ファイル1のシートを選択", common_sheets)
                        selected_sheets2 = [s for s in selected_sheets1]  # 同じシートを選択
                    else:
                        st.warning("共通するシートがありません")

                    # 全シートの比較結果を格納する配列
                    all_comparison_results = []
                    
                    # 各シートペアを比較
                    for sheet1, sheet2 in zip(selected_sheets1, selected_sheets2):
                        st.subheader(f"シートの比較: {sheet1} vs {sheet2}")
                        
                        try:
                            # Reset file pointers for reading Excel
                            file1.seek(0)
                            file2.seek(0)
                            
                            # Load sheet data
                            df1 = pd.read_excel(file1, sheet_name=sheet1)
                            df2 = pd.read_excel(file2, sheet_name=sheet2)
                            
                            if df1.empty or df2.empty:
                                st.warning(f"シート '{sheet1}' または '{sheet2}' にデータが存在しません。")
                                continue
                            
                            # Compare shapes
                            st.info(f"シート '{sheet1}' と '{sheet2}' の画像比較を開始...")
                            shapes1 = comparison.extract_shape_info(file1_path, sheet1)
                            st.write(f"ファイル1の画像数: {len(shapes1)}")
                            shapes2 = comparison.extract_shape_info(file2_path, sheet2)
                            st.write(f"ファイル2の画像数: {len(shapes2)}")
                            shape_differences = comparison.compare_shapes(shapes1, shapes2)
                            
                            # Compare data
                            comparison_result = comparison.compare_dataframes(df1, df2)
                            if not isinstance(comparison_result, dict):
                                st.error("比較結果の形式が不正です。")
                                continue
                            
                            # Add sheet names and shape differences to results
                            comparison_result['sheet1_name'] = sheet1
                            comparison_result['sheet2_name'] = sheet2
                            comparison_result['shape_differences'] = shape_differences
                            
                            all_comparison_results.append(comparison_result)
                            
                            # Display individual sheet comparison results
                            col1, col2 = st.columns(2)
                            
                            with col1:
                                st.markdown(f"### ファイル 1 - {sheet1}")
                                if 'df1' in comparison_result and not comparison_result['df1'].empty:
                                    grid1 = utils.create_grid(comparison_result['df1'], comparison_result.get('df1_styles', None))
                                else:
                                    st.warning("データを表示できません。")
                            
                            with col2:
                                st.markdown(f"### ファイル 2 - {sheet2}")
                                if 'df2' in comparison_result and not comparison_result['df2'].empty:
                                    grid2 = utils.create_grid(comparison_result['df2'], comparison_result.get('df2_styles', None))
                                else:
                                    st.warning("データを表示できません。")
                            
                            # Display shape differences for this sheet pair
                            if shape_differences:
                                st.subheader(f"図形の差分 ({sheet1} vs {sheet2})")
                                utils.display_shape_differences(shape_differences)
                            
                        except Exception as e:
                            st.error(f"シート '{sheet1}' と '{sheet2}' の比較中にエラーが発生しました: {str(e)}")
                            continue
                    
                    # 全シートの比較が完了した後、サマリーを表示
                    if all_comparison_results:
                        st.markdown("---")
                        st.subheader("全体の比較結果サマリー")
                        
                        # Create combined summary DataFrame
                        all_summary_data = []
                        
                        for result in all_comparison_results:
                            sheet1_name = result.get('sheet1_name', 'Unknown Sheet 1')
                            sheet2_name = result.get('sheet2_name', 'Unknown Sheet 2')
                            sheet_pair = f"{sheet1_name} → {sheet2_name}"
                            
                            # データの変更を処理
                            data_changes = []
                            if 'diff_summary' in result:
                                for diff in result['diff_summary'].to_dict('records'):
                                    change_info = {
                                        'シート名': sheet_pair,
                                        '変更タイプ': 'データ変更'
                                    }
                                    
                                    if diff['type'] == 'modified':
                                        change_info.update({
                                            'セル位置 (変更前)': get_excel_cell_reference(result['df1'].columns.get_loc(diff['column']), diff['row_index_old']),
                                            'セル位置 (変更後)': get_excel_cell_reference(result['df1'].columns.get_loc(diff['column']), diff['row_index_new']),
                                            '変更前の値': diff['value_old'],
                                            '変更後の値': diff['value_new']
                                        })
                                    else:
                                        df = result['df1'] if diff['type'] == 'deleted' else result['df2']
                                        range_ref = get_excel_range_reference(diff['row_index'], 0, len(df.columns) - 1)
                                        change_info.update({
                                            '変更タイプ': '行追加' if diff['type'] == 'added' else '行削除',
                                            'セル位置': f"{diff['row_index'] + 1}行目 ({range_ref})",
                                            '値': ' | '.join([f"{k}: {v}" for k, v in (diff['values'].items() if isinstance(diff['values'], dict) else {}).items() if pd.notna(v)])
                                        })
                                    
                                    data_changes.append(change_info)
                            
                            # 図形の変更を処理
                            shape_changes = []
                            if 'shape_differences' in result:
                                for shape_diff in result['shape_differences']:
                                    shape_info = {
                                        'シート名': sheet_pair,
                                        '変更タイプ': f'図形{shape_diff["type"]}'
                                    }
                                    
                                    if shape_diff['type'] == 'modified':
                                        old_shape = shape_diff['old_shape']
                                        new_shape = shape_diff['new_shape']
                                        shape_info.update({
                                            'セル位置 (変更前)': get_excel_cell_reference(old_shape['x'], old_shape['y']),
                                            'セル位置 (変更後)': get_excel_cell_reference(new_shape['x'], new_shape['y']),
                                            '変更前の値': f"Type: {old_shape['type']}, Text: {old_shape.get('text', '')}",
                                            '変更後の値': f"Type: {new_shape['type']}, Text: {new_shape.get('text', '')}"
                                        })
                                    else:
                                        shape = shape_diff.get('shape', {})
                                        shape_info.update({
                                            'セル位置': get_excel_cell_reference(shape['x'], shape['y']),
                                            '値': f"Type: {shape['type']}, Text: {shape.get('text', '')}"
                                        })
                                    
                                    shape_changes.append(shape_info)
                            
                            # シートごとの変更をまとめて追加
                            all_summary_data.extend(data_changes + shape_changes)
                            
                            # 元のデータを保存
                            result['df1'].to_excel(writer, sheet_name=f'F1_{sheet1_name[:26]}', index=False)
                            result['df2'].to_excel(writer, sheet_name=f'F2_{sheet2_name[:26]}', index=False)
                        
                        # サマリーシートの作成
                        if all_summary_data:
                            summary_df = pd.DataFrame(all_summary_data)
                            st.dataframe(summary_df)
                        else:
                            st.info("全シートで差分は検出されませんでした")
                        
                        # Export options for all sheets
                        st.markdown("---")
                        st.subheader("エクスポート")
                        utils.export_comparison(all_comparison_results, sheets1, sheets2)
                    
                except Exception as e:
                    st.error(f"比較処理中にエラーが発生しました: {str(e)}")
                    return
                
            except Exception as e:
                st.error(f"ファイルの処理中にエラーが発生しました: {str(e)}")
                return
        else:
            st.info("比較を開始するには、両方のExcelファイルをアップロードしてください")
    
    except Exception as e:
        st.error(f"予期せぬエラーが発生しました: {str(e)}")

if __name__ == "__main__":
    main()
